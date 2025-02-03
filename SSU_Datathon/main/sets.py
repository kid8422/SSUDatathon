from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from django.db import connection
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
import xgboost as xgb
from openpyxl import load_workbook
import csv
import io
import os
import re
import math
import json
from .models import Book
from django.db.models import Q

##############################################################
########################## 도서 정보 ##########################
###############################################################
@login_required(login_url='DB_login')
def download_book_data(request):
    if request.method == "POST":
        try: 
            with connection.cursor() as cursor:
                cursor.execute("SELECT ID, registration, get_course, DDC, ISBN, title, author, publisher, publication_year, location FROM book")
                book_data = cursor.fetchall()

            keys = ["도서ID", "등록일자", "수서방법", "분류코드", "ISBN",
                    "서명", "저자", "출판사", "출판연도", "소장위치"]
            
            # 각 값을 작은따옴표(')로 감싼 문자열로 변환
            transformed_data = [
                {keys[i]: f"'{str(row[i])}'" for i in range(len(keys))} 
                for row in book_data
            ]
            
            return JsonResponse({'success': True, 'data': transformed_data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def load_book_data(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # 요청 데이터 파싱
            page = int(data.get('page', 1))  # 선택된 데이터 가져오기
            pageSize = int(data.get('pageSize', 25))
            start = (page - 1) * pageSize
            #print(f"page : {page}, pageSize : {pageSize}, start : {start}")
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT ID, registration, get_course, DDC, ISBN, title, author, publisher, publication_year, location, `except` FROM book LIMIT {pageSize} OFFSET {start}")
                book_data = cursor.fetchall()
            return JsonResponse({'success': True, 'data': book_data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def load_book_max_page_len(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)  # 요청 데이터 파싱
            pageSize = int(body.get("pageSize", 25))
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM large_classification WHERE TAG = '전체'")
                book_data = cursor.fetchall()[0]
            total_count = sum(book_data[1:])
            return JsonResponse({'success': True, 'data': math.ceil(total_count / pageSize)})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

def data_check(data):
    if not re.match(r'^SS_\d{6}$', data['bookId']):
        print(repr(data['bookId']))  # 문자열 내부를 그대로 출력
        return "도서ID는 'SS_'로 시작하고 6개의 숫자로 구성되어야 합니다."
    
    try:
        datetime.strptime(data['regDate'], '%Y-%m-%d')
    except ValueError:
        return "등록일자는 'YYYY-MM-DD' 형식의 날짜여야 합니다."
    
    if not re.match(r'^\d{3}(\.\d+)?$', data['ddc']):
        print(repr(data['ddc']))  # 문자열 내부를 그대로 출력
        return "분류코드는 세 자리 숫자 또는 세 자리 숫자 뒤에 '.'과 숫자가 오는 형식이어야 합니다."
    
    if not re.match(r'^\d{4}$', data['pubYear']):
        return "출판연도는 4자리의 연도여야 합니다."
    
    # 검사 결과 문제 없음
    return "문제없음"

def data_preprocess(raw_data):
    data = {"ID" : raw_data["bookId"], 
            "registration" : raw_data["regDate"], 
            "get_course" : raw_data["getMethod"], 
            "DDC" : raw_data["ddc"], 
            "ISBN" : raw_data["isbn"], 
            "title" : raw_data["title"], 
            "author" : raw_data["author"], 
            "publisher" : raw_data["publisher"], 
            "publication_year" : raw_data["pubYear"],
            "location": raw_data["location"],
            "except": '1' if raw_data["checked"] else '0',
        }
    data["registration_year"] = int(raw_data["regDate"].split("-")[0])
    data["registration_month"] = str(int(raw_data["regDate"].split("-")[1]))
    data["large_code"] = raw_data["ddc"][0]
    data["middle_code"] = raw_data["ddc"][1]
    data["jaum"] = load_chosung(raw_data["title"])
    data["bulli"] = load_bulli(raw_data["title"])
    return data

# 초성, 중성, 종성 리스트
chosung = [
    "ㄱ", "ㄲ", "ㄴ", "ㄷ", "ㄸ", "ㄹ", "ㅁ", "ㅂ", "ㅃ", "ㅅ", "ㅆ",
    "ㅇ", "ㅈ", "ㅉ", "ㅊ", "ㅋ", "ㅌ", "ㅍ", "ㅎ"
]
jungsung = [
    "ㅏ", "ㅐ", "ㅑ", "ㅒ", "ㅓ", "ㅔ", "ㅕ", "ㅖ", "ㅗ", "ㅘ", "ㅙ",
    "ㅚ", "ㅛ", "ㅜ", "ㅝ", "ㅞ", "ㅟ", "ㅠ", "ㅡ", "ㅢ", "ㅣ"
]
jongsung = [
    "", "ㄱ", "ㄲ", "ㄳ", "ㄴ", "ㄵ", "ㄶ", "ㄷ", "ㄹ", "ㄺ", "ㄻ", "ㄼ",
    "ㄽ", "ㄾ", "ㄿ", "ㅀ", "ㅁ", "ㅂ", "ㅄ", "ㅅ", "ㅆ", "ㅇ", "ㅈ", "ㅊ",
    "ㅋ", "ㅌ", "ㅍ", "ㅎ"
]
    
def load_chosung(text):
    def extract_chosung(char):
        # 한글 유니코드 범위 확인
        if '가' <= char <= '힣':
            base = ord('가')
            char_code = ord(char) - base
            # 초성 계산
            cho = char_code // 588
            return chosung[cho]
        else:
            return char  # 한글이 아니면 그대로 반환
    # 입력 텍스트를 한 글자씩 처리하면서 띄어쓰기만 제거
    return ''.join(extract_chosung(char) if char != ' ' else '' for char in text)

def load_bulli(text):
    def decompose_korean(char):
        # 한글 유니코드 범위 확인
        if '가' <= char <= '힣':
            base = ord('가')
            char_code = ord(char) - base
            # 초성, 중성, 종성 분리
            cho = char_code // 588
            jung = (char_code % 588) // 28
            jong = char_code % 28
            # 초성, 중성, 종성 반환
            return chosung[cho] + jungsung[jung] + (jongsung[jong] if jong != 0 else "")
        else:
            return char  # 한글이 아니면 그대로 반환
    # 입력 텍스트를 한 글자씩 분리 후 처리
    #return ''.join(decompose_korean(char) for char in text)
    return ''.join(decompose_korean(char) if char != ' ' else '' for char in text)

def convert_excel_date(excel_date):
    dt = datetime.strptime(excel_date, '%Y-%m-%d')  # 문자열 → datetime 변환
    convertExcel_date = (dt - datetime(1899, 12, 30)).days  # 엑셀 날짜 변환

    base_date = datetime(1899, 12, 30)  # 엑셀의 날짜 오차 보정
    return (base_date + timedelta(days=convertExcel_date)).strftime('%Y-%m-%d')

@login_required(login_url='DB_login')
def save_add_book(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            stat = data_check(body)
            if stat == "문제없음":
                data = data_preprocess(body)
                print(f"""({data["ID"]}, {data["registration_year"]},{data["registration_month"]}, {data["get_course"]}, {data["DDC"]}, {data["ISBN"]}, {data["title"]}, {data["author"]}, {data["publisher"]}, {data["publication_year"]}, {data["location"]}, {data["large_code"]}, {data["middle_code"]}, {data["jaum"]}, {data["bulli"]}, {data["registration"]}, {data["except"]})""")
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            CALL insert_book(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (data["ID"], data["registration_year"], data["registration_month"], data["get_course"], data["DDC"], 
                            data["ISBN"], data["title"], data["author"], data["publisher"], data["publication_year"], data["location"], 
                            data["large_code"], data["middle_code"], data["jaum"], data["bulli"], data["registration"], data["except"]))
                    return JsonResponse({'success': True})
                except Exception as e:
                    if e.args[0] == 1062:  # MySQL 에러 코드 1062 (Duplicate entry)
                        return JsonResponse({'success': False, 'message': "이미 존재하는 도서ID입니다."}, status=400)
                    print(f"Error: {e}")
            else:
                return JsonResponse({'success': False, 'message': stat}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def save_book_file(request):
    if request.method == "POST":
        try:
            if 'file' not in request.FILES:
                return JsonResponse({"success": False, "message": "파일이 없습니다."}, status=400)
            
            uploaded_file = request.FILES['file']  # 업로드된 파일 가져오기

            # 2️⃣ JSON 데이터 읽기 (컬럼 매핑)
            extra_data = request.POST.get("extraData", "{}")  # JSON 문자열 가져오기
            extra_data = json.loads(extra_data)  # JSON 파싱

            if "cols" not in extra_data or "rows" not in extra_data:
                return JsonResponse({"success": False, "message": "컬럼 매핑 정보가 없습니다."}, status=400)
            
            col_mapping = {col: ord(col) - ord('A') for col in extra_data['cols']}

            # 3️⃣ 파일 확장자에 따라 처리 방식 결정
            file_ext = uploaded_file.name.split('.')[-1].lower()
            if file_ext in ['xls', 'xlsx']:
                # 엑셀 파일 처리 (openpyxl)
                wb = load_workbook(uploaded_file, data_only=True)
                sheet = wb.active  # 첫 번째 시트 사용
                rows = list(sheet.iter_rows(values_only=True))  # 모든 행을 리스트로 변환
            elif file_ext == 'csv':
                # CSV 파일 처리
                file_data = uploaded_file.read().decode('utf-8')
                csv_reader = csv.reader(io.StringIO(file_data))
                rows = list(csv_reader)  # 모든 행을 리스트로 변환
            else:
                return JsonResponse({"success": False, "message": "지원되지 않는 파일 형식입니다."}, status=400)

            # 4️⃣ Pandas DataFrame 생성
            df_columns = [
                "ID", "registration", "get_course", "DDC", "ISBN",
                "title", "author", "publisher", "publication_year", "location"
            ]
            df_data = []
            for row in rows:
                df_row = [row[col_mapping[col]] if col in col_mapping and col_mapping[col] < len(row) else None for col in extra_data['cols']]
                df_data.append(df_row)

            df = pd.DataFrame(df_data, columns=df_columns)
            # 모든 문자열 컬럼에서 작은 따옴표 제거
            df.at[0, 'ID'] = df.at[0, 'ID'].replace('\ufeff', '')
            df = df.applymap(lambda x: x.strip("'\"") if isinstance(x, str) else x)


            df["registration"] = df["registration"].apply(convert_excel_date)
            df["publication_year"] = df["publication_year"].astype(str)
            df["DDC"] = df["DDC"].astype(str)
            df["ISBN"] = df["ISBN"].astype(str)

            df['validation'] = df.apply(lambda row: data_check({
                'bookId': row['ID'],
                'regDate': row['registration'],
                'ddc': row['DDC'],
                'pubYear': row['publication_year']
            }), axis=1)

            invalid_rows = df[df['validation'] != "문제없음"]

            if invalid_rows.empty:
                processed_data = df.apply(lambda row: data_preprocess({
                    'bookId': row['ID'],
                    'regDate': row['registration'],
                    'getMethod': row['get_course'],
                    'ddc': row['DDC'],
                    'isbn': row['ISBN'],
                    'title': row['title'],
                    'author': row['author'],
                    'publisher': row['publisher'],
                    'pubYear': row['publication_year'],
                    'location': row['location'],
                    'checked': False,
                }), axis=1)
                df_processed = pd.DataFrame(processed_data.tolist())
                try:
                    with connection.cursor() as cursor:
                        insert_query = "CALL insert_book(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                        data_tuples = [
                            (
                                row["ID"], row["registration_year"], row["registration_month"], row["get_course"], row["DDC"],
                                row["ISBN"], row["title"], row["author"], row["publisher"], row["publication_year"],
                                row["location"], row["large_code"], row["middle_code"], row["jaum"], row["bulli"],
                                row["registration"], row["except"]
                            )
                            for _, row in df_processed.iterrows()
                        ]
                        cursor.executemany(insert_query, data_tuples)
                    connection.commit()
                    return JsonResponse({'success': True})
                except Exception as e:
                    connection.rollback()  # 오류 발생 시 롤백
                    print(f"Error: {e}")
                    if e.args[0] == 1062:  # MySQL 에러 코드 1062 (Duplicate entry)
                        duplicate_id = e.args[1].split("'")[1]  # "SS_000001" 추출
                        return JsonResponse({'success': False, 'message': f"'{duplicate_id}'은 이미 존재하는 도서ID입니다."}, status=400)
                    return JsonResponse({'success': False, 'message': f"알 수 없는 오류: {str(e)}"}, status=500)
            else:
                first_invalid_row = invalid_rows.iloc[0]  # 첫 번째 행 가져오기
                first_invalid_id = first_invalid_row["ID"]  # ID 컬럼 값
                first_invalid_validation = first_invalid_row["validation"]  # validation 컬럼 값
                text = f'{first_invalid_id}의 {first_invalid_validation} \n다시 확인해 주세요.'
                return JsonResponse({'success': False, 'message': text}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def edit_book(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            stat = data_check(body)
            if stat == "문제없음":
                data = data_preprocess(body)
                print(f"""({data["ID"]}, {data["registration_year"]},{data["registration_month"]}, {data["get_course"]}, {data["DDC"]}, {data["ISBN"]}, {data["title"]}, {data["author"]}, {data["publisher"]}, {data["publication_year"]}, {data["location"]}, {data["large_code"]}, {data["middle_code"]}, {data["jaum"]}, {data["bulli"]}, {data["registration"]}, {data["except"]})""")
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            CALL update_book(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (data["ID"], data["registration_year"], data["registration_month"], data["get_course"], data["DDC"], data["ISBN"], 
                            data["title"], data["author"], data["publisher"], data["publication_year"], data["location"], data["large_code"], 
                            data["middle_code"], data["jaum"], data["bulli"], data["registration"], data["except"]))
                    return JsonResponse({'success': True})
                except Exception as e:
                    if e.args[0] == 1062:  # MySQL 에러 코드 1062 (Duplicate entry)
                        return JsonResponse({'success': False, 'message': "이미 존재하는 도서ID입니다."}, status=400)
                    print(f"Error: {e}")
            else:
                return JsonResponse({'success': False, 'message': stat}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

##############################################################
########################## 대출 정보 ##########################
###############################################################
@login_required(login_url='DB_login')
def download_rent_data(request):
    if request.method == "POST":
        try: 
            with connection.cursor() as cursor:
                cursor.execute("SELECT ID, rent_date FROM rent")
                book_data = cursor.fetchall()

            keys = ["도서ID", "대출일시"]
            transformed_data = [dict(zip(keys, row)) for row in book_data]
            print(transformed_data)
            return JsonResponse({'success': True, 'data': transformed_data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def load_rent_data(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # 요청 데이터 파싱
            page = int(data.get('page', 1))  # 선택된 데이터 가져오기
            pageSize = int(data.get('pageSize', 25))
            order = int(data.get('order', 1))
            if order == 1:
                order_by = "DESC"
            else:
                order_by = "ASC"
            start = (page - 1) * pageSize
            #print(f"page : {page}, pageSize : {pageSize}, start : {start}")
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT rent_time, ID, title, author, publisher, DDC, location FROM rent_info ORDER BY rent_time {order_by} LIMIT {pageSize} OFFSET {start}")
                rent_data = cursor.fetchall()
            return JsonResponse({'success': True, 'data': rent_data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def load_rent_max_page_len(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)  # 요청 데이터 파싱
            pageSize = int(body.get("pageSize", 25))
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM year_month_count")
                rent_data = cursor.fetchall()
                total_sum = sum(sum(year_data[1:]) for year_data in rent_data)

            return JsonResponse({'success': True, 'data': math.ceil(total_sum / pageSize)})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

def convert_excel_datetime(excel_date_str):
    """
    excel_date_str: "YYYY-MM-DD HH:MM:SS" 형식의 문자열 (예: "2004-11-15 12:33:01")
    반환값: 엑셀 기준일(1899-12-30)로부터 days+fraction을 구하고,
           다시 1899-12-30에 더해 실제 날짜/시분초로 복원한 뒤
           "YYYY-MM-DD HH:MM:SS" 형식 문자열로 반환
    """

    # 1) 문자열을 datetime으로 파싱 (시분초 포함)
    dt = datetime.strptime(excel_date_str, '%Y-%m-%d %H:%M:%S')

    # 2) 엑셀 기준일(1899-12-30)과의 차이 계산
    #    - days + fraction(시분초에 따른 부분일) 형태의 float 값으로 엑셀 날짜 표현
    base_date = datetime(1899, 12, 30)
    diff = dt - base_date

    # diff.days는 정수 일수, diff.seconds는 0~86399 사이, diff.microseconds는 더 미세한 시간
    # 엑셀에서 시분초는 1일=1.0로 환산 -> 시분초를 일 단위 소수로 변환
    excel_float = diff.days + diff.seconds / 86400 + diff.microseconds / 86400000000

    # 3) 다시 엑셀 날짜(부동소수) -> datetime으로 복원
    restored_dt = base_date + timedelta(days=excel_float)

    # 4) "YYYY-MM-DD HH:MM:SS" 형식으로 포매팅하여 반환
    return restored_dt.strftime('%Y-%m-%d %H:%M:%S')

@login_required(login_url='DB_login')
def save_rent_file(request):
    if request.method == "POST":
        try:
            if 'file' not in request.FILES:
                return JsonResponse({"success": False, "message": "파일이 없습니다."}, status=400)
            
            uploaded_file = request.FILES['file']  # 업로드된 파일 가져오기

            # 2️⃣ JSON 데이터 읽기 (컬럼 매핑)
            extra_data = request.POST.get("extraData", "{}")  # JSON 문자열 가져오기
            extra_data = json.loads(extra_data)  # JSON 파싱

            if "cols" not in extra_data or "rows" not in extra_data:
                return JsonResponse({"success": False, "message": "컬럼 매핑 정보가 없습니다."}, status=400)
            
            col_mapping = {col: ord(col) - ord('A') for col in extra_data['cols']}

            # 3️⃣ 파일 확장자에 따라 처리 방식 결정
            file_ext = uploaded_file.name.split('.')[-1].lower()
            if file_ext in ['xls', 'xlsx']:
                # 엑셀 파일 처리 (openpyxl)
                wb = load_workbook(uploaded_file, data_only=True)
                sheet = wb.active  # 첫 번째 시트 사용
                rows = list(sheet.iter_rows(values_only=True))  # 모든 행을 리스트로 변환
            elif file_ext == 'csv':
                # CSV 파일 처리
                file_data = uploaded_file.read().decode('utf-8')
                csv_reader = csv.reader(io.StringIO(file_data))
                rows = list(csv_reader)  # 모든 행을 리스트로 변환
            else:
                return JsonResponse({"success": False, "message": "지원되지 않는 파일 형식입니다."}, status=400)

            df_columns = ["도서ID", "대출일시"]
            df_data = []
            for row in rows:
                df_row = [row[col_mapping[col]] if col in col_mapping and col_mapping[col] < len(row) else None for col in extra_data['cols']]
                df_data.append(df_row)

            df = pd.DataFrame(df_data, columns=df_columns)
            # 모든 문자열 컬럼에서 작은 따옴표 제거
            df.at[0, '도서ID'] = df.at[0, '도서ID'].replace('\ufeff', '')
            df = df.applymap(lambda x: x.strip("'\"") if isinstance(x, str) else x)

            
            df["대출일시"] = df["대출일시"].str.replace('T', ' ')
            df["대출일시"] = df["대출일시"].apply(convert_excel_datetime)


            try:
                with connection.cursor() as cursor:
                    insert_query = "CALL insert_rent(%s, %s)"
                    data_tuples = [
                        (
                            row["도서ID"], row["대출일시"]
                        )
                        for _, row in df.iterrows()
                    ]
                    cursor.executemany(insert_query, data_tuples)
                connection.commit()
                return JsonResponse({'success': True})
            except Exception as e:
                connection.rollback()  # 오류 발생 시 롤백
                print(f"Error: {e}")
                if e.args[0] == 1062:  # MySQL 에러 코드 1062 (Duplicate entry)
                    duplicate_id = e.args[1].split("'")[1]  # "SS_000001" 추출
                    return JsonResponse({'success': False, 'message': f"'{duplicate_id}'은 이미 존재하는 도서ID입니다."}, status=400)
                return JsonResponse({'success': False, 'message': f"알 수 없는 오류: {str(e)}"}, status=500)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

##############################################################
####################### 예외 도서 정보 ########################
###############################################################
@login_required(login_url='DB_login')
def except_download_book_data(request):
    if request.method == "POST":
        try: 
            with connection.cursor() as cursor:
                cursor.execute("SELECT ID, registration, get_course, DDC, ISBN, title, author, publisher, publication_year, location FROM book WHERE `except` = '1'")
                book_data = cursor.fetchall()

            keys = ["도서ID", "등록일자", "수서방법", "분류코드", "ISBN",
                "서명", "저자", "출판사", "출판연도", "소장위치"]
            transformed_data = [dict(zip(keys, row)) for row in book_data]
            return JsonResponse({'success': True, 'data': transformed_data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def except_load_book_data(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # 요청 데이터 파싱
            page = int(data.get('page', 1))  # 선택된 데이터 가져오기
            pageSize = int(data.get('pageSize', 25))
            start = (page - 1) * pageSize
            #print(f"page : {page}, pageSize : {pageSize}, start : {start}")
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT ID, registration, get_course, DDC, ISBN, title, author, publisher, publication_year, location, `except` FROM book WHERE `except` = '1' LIMIT {pageSize} OFFSET {start}")
                book_data = cursor.fetchall()
            return JsonResponse({'success': True, 'data': book_data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def except_load_book_max_page_len(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)  # 요청 데이터 파싱
            pageSize = int(body.get("pageSize", 25))
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT count(*) FROM book WHERE `except` = '1'")
                book_data = cursor.fetchone()[0]
            print(book_data)
            return JsonResponse({'success': True, 'data': math.ceil(book_data / pageSize)})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def save_except_book_file(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            col_mapping = {col: ord(col) - ord('A') for col in body['cols']}
            df_columns = ["ID"]
            df_data = []
            for row in body['rows']:
                df_row = [row[col_mapping[col]] for col in body['cols']]
                df_data.append(df_row)
            df = pd.DataFrame(df_data, columns=df_columns)


            df['validation'] = df['ID'].apply(lambda x: "문제없음" if re.match(r'^SS_\d{6}$', x) else "도서ID는 'SS_'로 시작하고 6개의 숫자로 구성되어야 합니다.")

            invalid_rows = df[df['validation'] != "문제없음"]

            if invalid_rows.empty:
                try:
                    with connection.cursor() as cursor:
                        insert_query = """
                            UPDATE book SET `except` = '1' WHERE ID = %s
                        """
                        data_tuples = [
                            (
                                row["ID"]
                            )
                            for _, row in df.iterrows()
                        ]
                        cursor.executemany(insert_query, data_tuples)
                    connection.commit()
                    return JsonResponse({'success': True})
                except Exception as e:
                    connection.rollback()  # 오류 발생 시 롤백
                    print(f"Error: {e}")
                    if e.args[0] == 1062:  # MySQL 에러 코드 1062 (Duplicate entry)
                        duplicate_id = e.args[1].split("'")[1]  # "SS_000001" 추출
                        return JsonResponse({'success': False, 'message': f"'{duplicate_id}'은 이미 존재하는 도서ID입니다."}, status=400)
                    return JsonResponse({'success': False, 'message': f"알 수 없는 오류: {str(e)}"}, status=500)
            else:
                first_invalid_row = invalid_rows.iloc[0]  # 첫 번째 행 가져오기
                first_invalid_id = first_invalid_row["ID"]  # ID 컬럼 값
                first_invalid_validation = first_invalid_row["validation"]  # validation 컬럼 값
                text = f'{first_invalid_id}의 {first_invalid_validation} \n다시 확인해 주세요.'
                return JsonResponse({'success': False, 'message': text}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

##############################################################
####################### 데이터 전처리 #########################
###############################################################

@login_required(login_url='DB_login')
def load_db_update(request):
    if request.method == "POST":
        try:
            db = ["book", "rent", "large_classification", "middle_classification", 
                  "rent_count", "year_month_count", "year_month_count_detail", 
                  "ISBN_rent_count", "None_ISBN_rent_count","recent_rent", ]
            data = {}

            with connection.cursor() as cursor:
                for table in db:
                    cursor.execute(f"SELECT `timestamp` FROM timestamp WHERE `table` = '{table}'")
                    raw_data = cursor.fetchall()[0]
                    if raw_data[0] == None:
                        d = "변동 내역 없음"
                    else:
                        d = raw_data[0].strftime("%Y년 %m월 %d일 %H시 %M분 %S초").replace(" 0", " ")
                    data[table] = d
            
            return JsonResponse({'success': True, 'data': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

@login_required(login_url='DB_login')
def preprocessing_db(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # 요청 데이터 파싱
            
            option = data.get('data')  # 선택된 데이터 가져오기

            error = []

            if 'ISBN 제공 도서 대여 정보' in option:
                try: 
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT ID, rent_date, ISBN FROM rent NATURAL JOIN book WHERE rent.ID = book.ID")
                        raw_data = cursor.fetchall()

                        ISBN_df = pd.DataFrame(raw_data, columns=['ID', 'rent_date', 'ISBN'])
                        ISBN_df['rent_year'] = pd.to_datetime(ISBN_df['rent_date']).dt.year
                        ISBN_df = ISBN_df[ISBN_df['ISBN'] != '0']

                        result_df = ISBN_df.groupby('ISBN').agg(
                            도서ID개수=('ID', lambda x: x.nunique()),  # 고유 도서 ID 개수
                        ).reset_index()

                        years = list(range(2004, 2025))
                        for year in years:
                            result_df[year] = 0  # 기본값 0 설정
                        year_counts = ISBN_df.groupby(['ISBN', 'rent_year']).size().unstack(fill_value=0)

                        for year in years:
                            if year in year_counts.columns:
                                result_df[year] = result_df['ISBN'].map(year_counts[year])
                        result_df = result_df.astype(object)
                except Exception as e:
                    error.append(str(e))

                try:
                    with connection.cursor() as cursor:
                        insert_query = """
                            INSERT INTO ISBN_rent_count 
                            (ISBN, ID_count, `2004`, `2005`, `2006`, `2007`, `2008`, `2009`, `2010`, `2011`, `2012`, `2013`, 
                            `2014`, `2015`, `2016`, `2017`, `2018`, `2019`, `2020`, `2021`, `2022`, `2023`, `2024`) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE 
                            ID_count = VALUES(ID_count),
                            `2004` = VALUES(`2004`), `2005` = VALUES(`2005`), `2006` = VALUES(`2006`), 
                            `2007` = VALUES(`2007`), `2008` = VALUES(`2008`), `2009` = VALUES(`2009`), 
                            `2010` = VALUES(`2010`), `2011` = VALUES(`2011`), `2012` = VALUES(`2012`), 
                            `2013` = VALUES(`2013`), `2014` = VALUES(`2014`), `2015` = VALUES(`2015`), 
                            `2016` = VALUES(`2016`), `2017` = VALUES(`2017`), `2018` = VALUES(`2018`), 
                            `2019` = VALUES(`2019`), `2020` = VALUES(`2020`), `2021` = VALUES(`2021`), 
                            `2022` = VALUES(`2022`), `2023` = VALUES(`2023`), `2024` = VALUES(`2024`)
                        """
                        data_tuples = [
                            (
                                result_df.iloc[i, 0], result_df.iloc[i, 1], result_df.iloc[i, 2], result_df.iloc[i, 3], result_df.iloc[i, 4], result_df.iloc[i, 5], 
                                result_df.iloc[i, 6], result_df.iloc[i, 7], result_df.iloc[i, 8], result_df.iloc[i, 9], result_df.iloc[i, 10], result_df.iloc[i, 11], 
                                result_df.iloc[i, 12], result_df.iloc[i, 13], result_df.iloc[i, 14], result_df.iloc[i, 15], result_df.iloc[i, 16], result_df.iloc[i, 17], 
                                result_df.iloc[i, 18], result_df.iloc[i, 19], result_df.iloc[i, 20], result_df.iloc[i, 21], result_df.iloc[i, 22]
                            )
                            for i in range(len(result_df))
                        ]
                        cursor.executemany(insert_query, data_tuples)
                    connection.commit()
                except Exception as e:
                    connection.rollback()  # 오류 발생 시 롤백
                    error.append(str(e))


            if 'ISBN 미제공 도서 대여 정보' in option:
                try: 
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT ID, rent_date, ISBN, title, author, publisher FROM rent NATURAL JOIN book WHERE rent.ID = book.ID")
                        raw_data = cursor.fetchall()

                        ISBN_df = pd.DataFrame(raw_data, columns=['ID', 'rent_date', 'ISBN', '제목', '저자', '출판사'])
                        ISBN_df['rent_year'] = pd.to_datetime(ISBN_df['rent_date']).dt.year
                        ISBN_df = ISBN_df[ISBN_df['ISBN'] == '0']

                        result_df = ISBN_df.groupby(['제목', '저자', '출판사']).agg(
                            도서ID개수=('ID', lambda x: x.nunique()),  # 고유 도서 ID 개수
                        ).reset_index()

                        years = list(range(2004, 2025))
                        for year in years:
                            result_df[year] = 0  # 기본값 0 설정
                        year_counts = ISBN_df.groupby(['제목', '저자', '출판사', 'rent_year']).size().unstack(fill_value=0)

                        for year in years:
                            if year in year_counts.columns:
                                result_df[year] = result_df.set_index(['제목', '저자', '출판사']).index.map(year_counts[year])
                        result_df = result_df.astype(object)

                except Exception as e:
                    error.append(str(e))

                try:
                    with connection.cursor() as cursor:
                        insert_query = """
                            INSERT INTO None_ISBN_rent_count 
                            (title, author, publisher, ID_count, `2004`, `2005`, `2006`, `2007`, `2008`, `2009`, `2010`, `2011`, `2012`, `2013`, 
                            `2014`, `2015`, `2016`, `2017`, `2018`, `2019`, `2020`, `2021`, `2022`, `2023`, `2024`) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE 
                            ID_count = VALUES(ID_count),
                            `2004` = VALUES(`2004`), `2005` = VALUES(`2005`), `2006` = VALUES(`2006`), 
                            `2007` = VALUES(`2007`), `2008` = VALUES(`2008`), `2009` = VALUES(`2009`), 
                            `2010` = VALUES(`2010`), `2011` = VALUES(`2011`), `2012` = VALUES(`2012`), 
                            `2013` = VALUES(`2013`), `2014` = VALUES(`2014`), `2015` = VALUES(`2015`), 
                            `2016` = VALUES(`2016`), `2017` = VALUES(`2017`), `2018` = VALUES(`2018`), 
                            `2019` = VALUES(`2019`), `2020` = VALUES(`2020`), `2021` = VALUES(`2021`), 
                            `2022` = VALUES(`2022`), `2023` = VALUES(`2023`), `2024` = VALUES(`2024`);
                        """
                        data_tuples = [
                            (
                                result_df.iloc[i, 0], result_df.iloc[i, 1], result_df.iloc[i, 2], result_df.iloc[i, 3], result_df.iloc[i, 4], 
                                result_df.iloc[i, 5], result_df.iloc[i, 6], result_df.iloc[i, 7], result_df.iloc[i, 8], result_df.iloc[i, 9], 
                                result_df.iloc[i, 10], result_df.iloc[i, 11], result_df.iloc[i, 12], result_df.iloc[i, 13], result_df.iloc[i, 14], 
                                result_df.iloc[i, 15], result_df.iloc[i, 16], result_df.iloc[i, 17], result_df.iloc[i, 18], result_df.iloc[i, 19], 
                                result_df.iloc[i, 20], result_df.iloc[i, 21], result_df.iloc[i, 22], result_df.iloc[i, 23], result_df.iloc[i, 24]
                            )
                            for i in range(len(result_df))
                        ]
                        cursor.executemany(insert_query, data_tuples)
                    connection.commit()
                except Exception as e:
                    connection.rollback()  # 오류 발생 시 롤백
                    error.append(str(e))


            if '최근 대여 날짜 정보' in option:
                try: 
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT ID, registration FROM book")
                        book_rows = cursor.fetchall()

                        book_df = pd.DataFrame(book_rows, columns=['ID', '등록날짜'])
                        book_df['등록날짜'] = pd.to_datetime(book_df['등록날짜'])

                        cursor.execute("SELECT ID, rent_date FROM rent")
                        rent_rows = cursor.fetchall()

                        rent_df = pd.DataFrame(rent_rows, columns=['ID', '대여날짜'])
                        rent_df['대여날짜'] = pd.to_datetime(rent_df['대여날짜'])

                        latest_rent = rent_df.groupby('ID')['대여날짜'].max()

                        default_date = latest_rent.max()

                        book_df['Delta'] = None  # Delta 컬럼 생성 (초기값 None)
                        book_df.set_index('ID', inplace=True)

                        for book_id, last_rent_date in latest_rent.items():
                            book_df.at[book_id, 'Delta'] = (default_date - last_rent_date).days

                        book_df['Delta'] = book_df['Delta'].fillna((default_date - book_df['등록날짜']).dt.days)
                        book_df = book_df.infer_objects(copy=False)
                        book_df['Delta'] = book_df['Delta'].astype(int)

                        book_df.reset_index(inplace=True)
                        book_df['Delta'] = book_df['Delta'].astype(int)

                        book_df['Delta'] = book_df['Delta'].apply(lambda x: max(x, 0))
                except Exception as e:
                    error.append(str(e))
                
                try:
                    with connection.cursor() as cursor:
                        insert_query = """
                            INSERT INTO recent_rent (ID, duration) 
                            VALUES (%s, %s)
                            ON DUPLICATE KEY UPDATE 
                            duration = VALUES(duration);
                        """
                        data_tuples = [
                            (
                                book_df.iloc[i, 0], book_df.iloc[i, 2],
                            )
                            for i in range(len(book_df))
                        ]
                        cursor.executemany(insert_query, data_tuples)
                    connection.commit()
                except Exception as e:
                    connection.rollback()  # 오류 발생 시 롤백
                    error.append(str(e))

            if len(error) == 0:
                return JsonResponse({'success': True, 'data': data})
            else:
                return JsonResponse({'success': False, 'message': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)